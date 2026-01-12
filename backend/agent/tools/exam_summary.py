"""
Exam Result Summary Tool
Summarizes and exports exam results by exam code.
"""

import json
import datetime
from typing import Type, List, Dict, Any

from langchain.tools import BaseTool
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font
import pyodbc
import yagmail

from .base import check_role, get_role, format_permission_error, logger
from ...config import settings

__all__ = ["ExamResultSummaryTool", "ExamSummaryInput"]


class ExamSummaryInput(BaseModel):
    """Input schema for exam summary tool"""
    exam_code: str = Field(
        description="Mã đề thi cần tổng hợp (ví dụ: 000)"
    )


class ExamResultSummaryTool(BaseTool):
    """
    Tool tổng hợp kết quả bài thi theo mã đề.
    
    Features:
    - Query results from database
    - Calculate statistics (average, min, max)
    - Export to Excel
    - Send email with results
    """

    name: str = "summarize_exam_results"
    description: str = """
    Tổng hợp kết quả chấm điểm từ bảng FinalExamResult theo mã đề thi.

    Sử dụng khi người dùng yêu cầu:
    - Tổng hợp kết quả mã đề
    - Thống kê điểm bài thi
    - Xem kết quả toàn bộ sinh viên theo mã đề

    Tool sẽ:
    1. Query các cột cần thiết:
       - student_id, name, email, exam_code, score
    2. Trả về:
       - Danh sách sinh viên
       - Điểm trung bình, cao nhất, thấp nhất
       - Đánh giá tổng quan kết quả bài thi
       - File Excel tổng hợp
       - Gửi email file Excel đến địa chỉ cố định
    """
    args_schema: Type[BaseModel] = ExamSummaryInput

    def _run(self, exam_code: str) -> str:
        """
        Execute exam result summary.
        
        Args:
            exam_code: The exam code to summarize
            
        Returns:
            JSON string with summary data or error
        """
        try:
            # Check permission
            if not check_role("teacher"):
                return json.dumps(
                    format_permission_error("teacher"),
                    ensure_ascii=False, 
                    indent=2
                )
            
            logger.info(f"Summarizing results for exam_code={exam_code}")

            # Query database
            conn = pyodbc.connect(settings.SQL_SERVER_CONN_STR)
            cursor = conn.cursor()

            sql = """
            SELECT
                s.student_id AS student_id,
                s.full_name,
                s.email,
                fr.exam_code,
                fr.score
            FROM dbo.final_results fr
            JOIN dbo.students s
                ON s.student_id = fr.student_id
            WHERE fr.exam_code = ?
            ORDER BY fr.score DESC
            """

            cursor.execute(sql, exam_code)
            rows = cursor.fetchall()

            if not rows:
                cursor.close()
                conn.close()
                return json.dumps({
                    "exam_code": exam_code,
                    "message": "Không tìm thấy kết quả cho mã đề này"
                }, ensure_ascii=False, indent=2)

            # Process results
            results: List[Dict[str, Any]] = []
            scores: List[float] = []

            for r in rows:
                score = float(r.score)
                scores.append(score)

                results.append({
                    "student_id": r.student_id,
                    "full_name": r.full_name,
                    "email": r.email,
                    "exam_code": r.exam_code,
                    "score": score,
                    "evaluation": self._evaluate_score(score)
                })

            # Calculate summary statistics
            summary = {
                "total_students": len(scores),
                "average_score": round(sum(scores) / len(scores), 2),
                "max_score": max(scores),
                "min_score": min(scores),
            }

            assessment = self._overall_assessment(summary["average_score"])

            # Export to Excel
            excel_file = self._export_to_excel(
                exam_code=exam_code,
                summary=summary,
                results=results
            )

            # Send email
            sent = self._send_excel_email(
                file_path=excel_file,
                to_email=settings.EMAIL_RECEIVER,
                subject=f"Kết quả tổng hợp mã đề {exam_code}",
                body=f"Đính kèm file Excel tổng hợp kết quả bài thi mã đề {exam_code}."
            )
            
            if sent:
                logger.info(f"Excel file sent to {settings.EMAIL_RECEIVER}")
            else:
                logger.warning(f"Could not send Excel file to {settings.EMAIL_RECEIVER}")

            cursor.close()
            conn.close()

            return json.dumps({
                "exam_code": exam_code,
                "summary": summary,
                "overall_assessment": assessment,
                "results": results,
                "excel_file": excel_file,
                "email_sent": sent
            }, ensure_ascii=False, indent=2)

        except Exception as e:
            logger.error(f"Error summarizing exam results: {str(e)}")
            logger.exception(e)
            return json.dumps({
                "error": str(e),
                "type": type(e).__name__
            }, ensure_ascii=False, indent=2)

    async def _arun(self, exam_code: str) -> str:
        """Execute tool asynchronously"""
        return self._run(exam_code)

    def _evaluate_score(self, score: float) -> str:
        """
        Evaluate individual student score.
        
        Args:
            score: Student's score
            
        Returns:
            Evaluation string
        """
        if score >= 8.5:
            return "Xuất sắc"
        elif score >= 7.0:
            return "Tốt"
        elif score >= 5.0:
            return "Đạt"
        else:
            return "Chưa đạt"

    def _overall_assessment(self, avg_score: float) -> str:
        """
        Provide overall assessment based on average score.
        
        Args:
            avg_score: Class average score
            
        Returns:
            Assessment string
        """
        if avg_score >= 8.0:
            return "Kết quả bài thi rất tốt, đa số sinh viên nắm vững kiến thức."
        elif avg_score >= 6.5:
            return "Kết quả bài thi khá tốt, còn một số điểm cần cải thiện."
        elif avg_score >= 5.0:
            return "Kết quả ở mức trung bình, nhiều sinh viên còn hổng kiến thức."
        else:
            return "Kết quả thấp, cần xem lại đề thi hoặc phương pháp giảng dạy."
    
    def _send_excel_email(
        self, 
        file_path: str, 
        to_email: str, 
        subject: str, 
        body: str
    ) -> bool:
        """
        Send Excel file via email.
        
        Args:
            file_path: Path to Excel file
            to_email: Recipient email address
            subject: Email subject
            body: Email body
            
        Returns:
            True if sent successfully, False otherwise
        """
        try:
            yag = yagmail.SMTP(
                user=settings.EMAIL_USER, 
                password=settings.EMAIL_PASSWORD
            )
            yag.send(
                to=to_email,
                subject=subject,
                contents=body,
                attachments=file_path
            )
            logger.info(f"Excel file sent to {to_email}")
            return True
        except Exception as e:
            logger.error(f"Failed to send email: {str(e)}")
            return False    

    def _export_to_excel(
        self,
        exam_code: str,
        summary: Dict[str, Any],
        results: List[Dict[str, Any]]
    ) -> str:
        """
        Export results to Excel file.
        
        Args:
            exam_code: Exam code
            summary: Summary statistics
            results: List of student results
            
        Returns:
            Path to created Excel file
        """
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary["A1"] = "Exam Code"
        ws_summary["B1"] = exam_code

        ws_summary["A3"] = "Total Students"
        ws_summary["B3"] = summary["total_students"]

        ws_summary["A4"] = "Average Score"
        ws_summary["B4"] = summary["average_score"]

        ws_summary["A5"] = "Max Score"
        ws_summary["B5"] = summary["max_score"]

        ws_summary["A6"] = "Min Score"
        ws_summary["B6"] = summary["min_score"]

        for cell in ["A1", "A3", "A4", "A5", "A6"]:
            ws_summary[cell].font = Font(bold=True)

        # Results sheet
        ws_results = wb.create_sheet("Results")

        headers = [
            "Student ID",
            "Name",
            "Email",
            "Exam Code",
            "Score",
            "Evaluation"
        ]
        ws_results.append(headers)

        for col in range(1, len(headers) + 1):
            ws_results.cell(row=1, column=col).font = Font(bold=True)

        for r in results:
            ws_results.append([
                r["student_id"],
                r["full_name"],
                r["email"],
                r["exam_code"],
                r["score"],
                r["evaluation"]
            ])

        # Auto-adjust column widths
        for col in ws_results.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_results.column_dimensions[col[0].column_letter].width = max_len + 2

        # Save file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"exam_summary_{exam_code}_{timestamp}.xlsx"

        output_dir = settings.PROJECT_ROOT / "exports"
        output_dir.mkdir(parents=True, exist_ok=True)

        file_path = output_dir / filename
        wb.save(file_path)

        return str(file_path)
