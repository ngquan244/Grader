"""
Grading API routes
Handles exam grading and result summary
"""
import json
import datetime
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional
from fastapi import APIRouter, HTTPException

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from backend.schemas import GradingRequest, GradingResponse, GradingResult, GradingSummary
from backend.config import settings
from src.config import Config

try:
    import pyodbc
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import yagmail
except ImportError:
    pyodbc = None
    Workbook = None
    yagmail = None

router = APIRouter()


def check_role(role_required: str) -> bool:
    """Check if current role matches required role"""
    actual_role = Config.get_role()
    return (actual_role or "").lower() == role_required.lower()


@router.post("/execute")
async def execute_grading():
    """
    Execute the grading notebook to process uploaded exam images
    """
    try:
        # Import the notebook tool
        from src.notebook_tool import get_notebook_tool
        
        notebook_tool = get_notebook_tool()
        result = notebook_tool._run()
        
        return {
            "success": True,
            "result": json.loads(result) if isinstance(result, str) else result
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/summary", response_model=GradingResponse)
async def summarize_exam_results(request: GradingRequest):
    """
    Summarize exam results by exam code
    """
    if not check_role("teacher"):
        raise HTTPException(
            status_code=403,
            detail="Chỉ giáo viên mới có quyền xem tổng hợp kết quả"
        )
    
    if pyodbc is None:
        raise HTTPException(
            status_code=500,
            detail="Module pyodbc chưa được cài đặt"
        )
    
    try:
        conn = pyodbc.connect(settings.SQL_SERVER_CONN_STR)
        cursor = conn.cursor()
        
        sql = """
        SELECT
            s.student_id AS student_id,
            s.full_name,
            s.email,
            fr.exam_code,
            fr.score
        FROM dbo.final_results fr
        JOIN dbo.students s
            ON s.student_id = fr.student_id
        WHERE fr.exam_code = ?
        ORDER BY fr.score DESC
        """
        
        cursor.execute(sql, request.exam_code)
        rows = cursor.fetchall()
        
        if not rows:
            return GradingResponse(
                success=False,
                exam_code=request.exam_code or "",
                error="Không tìm thấy kết quả cho mã đề này"
            )
        
        results: List[GradingResult] = []
        scores: List[float] = []
        
        for r in rows:
            score = float(r.score)
            scores.append(score)
            
            results.append(GradingResult(
                student_id=r.student_id,
                full_name=r.full_name,
                email=r.email,
                exam_code=r.exam_code,
                score=score,
                evaluation=evaluate_score(score)
            ))
        
        summary = GradingSummary(
            total_students=len(scores),
            average_score=round(sum(scores) / len(scores), 2),
            max_score=max(scores),
            min_score=min(scores)
        )
        
        assessment = overall_assessment(summary.average_score)
        
        # Export to Excel
        excel_file = export_to_excel(
            exam_code=request.exam_code,
            summary=summary,
            results=results
        )
        
        # Send email
        send_excel_email(
            file_path=excel_file,
            to_email=settings.EMAIL_RECEIVER,
            subject=f"Kết quả tổng hợp mã đề {request.exam_code}",
            body=f"Đính kèm file Excel tổng hợp kết quả bài thi mã đề {request.exam_code}."
        )
        
        cursor.close()
        conn.close()
        
        return GradingResponse(
            success=True,
            exam_code=request.exam_code or "",
            summary=summary,
            overall_assessment=assessment,
            results=results,
            excel_file=excel_file
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/results")
async def get_all_results():
    """
    Get all grading results from JSON file
    """
    try:
        result_file = settings.PROJECT_ROOT / "final_result.json"
        if not result_file.exists():
            return {"success": False, "results": [], "message": "Chưa có kết quả chấm điểm"}
        
        with open(result_file, 'r', encoding='utf-8') as f:
            results = json.load(f)
        
        return {"success": True, "results": results}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def evaluate_score(score: float) -> str:
    """Evaluate score and return assessment"""
    if score >= 8.5:
        return "Xuất sắc"
    elif score >= 7.0:
        return "Tốt"
    elif score >= 5.0:
        return "Đạt"
    else:
        return "Chưa đạt"


def overall_assessment(avg_score: float) -> str:
    """Generate overall assessment based on average score"""
    if avg_score >= 8.0:
        return "Kết quả bài thi rất tốt, đa số sinh viên nắm vững kiến thức."
    elif avg_score >= 6.5:
        return "Kết quả bài thi khá tốt, còn một số điểm cần cải thiện."
    elif avg_score >= 5.0:
        return "Kết quả ở mức trung bình, nhiều sinh viên còn hổng kiến thức."
    else:
        return "Kết quả thấp, cần xem lại đề thi hoặc phương pháp giảng dạy."


def export_to_excel(exam_code: str, summary: GradingSummary, results: List[GradingResult]) -> str:
    """Export results to Excel file"""
    if Workbook is None:
        return ""
    
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary["A1"] = "Exam Code"
    ws_summary["B1"] = exam_code
    ws_summary["A3"] = "Total Students"
    ws_summary["B3"] = summary.total_students
    ws_summary["A4"] = "Average Score"
    ws_summary["B4"] = summary.average_score
    ws_summary["A5"] = "Max Score"
    ws_summary["B5"] = summary.max_score
    ws_summary["A6"] = "Min Score"
    ws_summary["B6"] = summary.min_score
    
    for cell in ["A1", "A3", "A4", "A5", "A6"]:
        ws_summary[cell].font = Font(bold=True)
    
    # Results sheet
    ws_results = wb.create_sheet("Results")
    
    headers = ["Student ID", "Name", "Email", "Exam Code", "Score", "Evaluation"]
    ws_results.append(headers)
    
    for col in range(1, len(headers) + 1):
        ws_results.cell(row=1, column=col).font = Font(bold=True)
    
    for r in results:
        ws_results.append([
            r.student_id,
            r.full_name,
            r.email,
            r.exam_code,
            r.score,
            r.evaluation
        ])
    
    # Auto-adjust column width
    for col in ws_results.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_results.column_dimensions[col[0].column_letter].width = max_len + 2
    
    # Save file
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"exam_summary_{exam_code}_{timestamp}.xlsx"
    
    file_path = settings.EXPORTS_DIR / filename
    wb.save(file_path)
    
    return str(file_path)


def send_excel_email(file_path: str, to_email: str, subject: str, body: str) -> bool:
    """Send Excel file via email"""
    if yagmail is None or not file_path:
        return False
    
    try:
        yag = yagmail.SMTP(user=settings.EMAIL_USER, password=settings.EMAIL_PASSWORD)
        yag.send(
            to=to_email,
            subject=subject,
            contents=body,
            attachments=file_path
        )
        return True
    except Exception:
        return False
