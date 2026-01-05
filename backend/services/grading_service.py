"""
Grading Service
Handles exam grading and result management
"""
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from backend.config import settings
from backend.core import ForbiddenException, NotFoundException, DatabaseException
from backend.utils import ensure_directory, evaluate_score

logger = logging.getLogger(__name__)

# Optional imports for database and email
try:
    import pyodbc
    PYODBC_AVAILABLE = True
except ImportError:
    pyodbc = None
    PYODBC_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = None
    OPENPYXL_AVAILABLE = False

try:
    import yagmail
    YAGMAIL_AVAILABLE = True
except ImportError:
    yagmail = None
    YAGMAIL_AVAILABLE = False


class GradingService:
    """Service for exam grading and result management"""
    
    def __init__(self):
        self.exports_dir = settings.EXPORTS_DIR
        self.results_file = settings.PROJECT_ROOT / "final_result.json"
        ensure_directory(self.exports_dir)
    
    def get_results_from_json(self) -> List[Dict[str, Any]]:
        """Get grading results from JSON file"""
        if not self.results_file.exists():
            return []
        
        with open(self.results_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def get_results_by_exam_code(self, exam_code: str) -> Dict[str, Any]:
        """Get grading results from database by exam code"""
        if not PYODBC_AVAILABLE:
            raise DatabaseException("query", "pyodbc not installed")
        
        try:
            conn = pyodbc.connect(settings.SQL_SERVER_CONN_STR)
            cursor = conn.cursor()
            
            sql = """
            SELECT
                s.student_id,
                s.full_name,
                s.email,
                fr.exam_code,
                fr.score
            FROM dbo.final_results fr
            JOIN dbo.students s ON s.student_id = fr.student_id
            WHERE fr.exam_code = ?
            ORDER BY fr.score DESC
            """
            
            cursor.execute(sql, exam_code)
            rows = cursor.fetchall()
            
            if not rows:
                raise NotFoundException("Kết quả mã đề", exam_code)
            
            results = []
            scores = []
            
            for r in rows:
                score = float(r.score)
                scores.append(score)
                results.append({
                    "student_id": r.student_id,
                    "full_name": r.full_name,
                    "email": r.email,
                    "exam_code": r.exam_code,
                    "score": score,
                    "evaluation": evaluate_score(score)
                })
            
            summary = {
                "total_students": len(scores),
                "average_score": round(sum(scores) / len(scores), 2),
                "max_score": max(scores),
                "min_score": min(scores)
            }
            
            cursor.close()
            conn.close()
            
            return {
                "exam_code": exam_code,
                "summary": summary,
                "overall_assessment": self._get_overall_assessment(summary["average_score"]),
                "results": results
            }
            
        except NotFoundException:
            raise
        except Exception as e:
            logger.error(f"Database error: {e}")
            raise DatabaseException("query results", str(e))
    
    def export_to_excel(
        self,
        exam_code: str,
        summary: Dict[str, Any],
        results: List[Dict[str, Any]]
    ) -> str:
        """Export results to Excel file"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, skipping Excel export")
            return ""
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        headers = [("A1", "Exam Code"), ("A3", "Total Students"),
                   ("A4", "Average Score"), ("A5", "Max Score"), ("A6", "Min Score")]
        values = [("B1", exam_code), ("B3", summary["total_students"]),
                  ("B4", summary["average_score"]), ("B5", summary["max_score"]),
                  ("B6", summary["min_score"])]
        
        for cell, val in headers:
            ws_summary[cell] = val
            ws_summary[cell].font = Font(bold=True)
        
        for cell, val in values:
            ws_summary[cell] = val
        
        # Results sheet
        ws_results = wb.create_sheet("Results")
        headers = ["Student ID", "Name", "Email", "Exam Code", "Score", "Evaluation"]
        ws_results.append(headers)
        
        for col in range(1, len(headers) + 1):
            ws_results.cell(row=1, column=col).font = Font(bold=True)
        
        for r in results:
            ws_results.append([
                r["student_id"], r["full_name"], r["email"],
                r["exam_code"], r["score"], r["evaluation"]
            ])
        
        # Auto-adjust column width
        for col in ws_results.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_results.column_dimensions[col[0].column_letter].width = max_len + 2
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"exam_summary_{exam_code}_{timestamp}.xlsx"
        file_path = self.exports_dir / filename
        wb.save(file_path)
        
        logger.info(f"Exported to Excel: {filename}")
        return str(file_path)
    
    def send_email(
        self,
        to_email: str,
        subject: str,
        body: str,
        attachment: Optional[str] = None
    ) -> bool:
        """Send email with optional attachment"""
        if not YAGMAIL_AVAILABLE:
            logger.warning("yagmail not available, skipping email")
            return False
        
        try:
            yag = yagmail.SMTP(
                user=settings.EMAIL_USER,
                password=settings.EMAIL_PASSWORD
            )
            
            kwargs = {"to": to_email, "subject": subject, "contents": body}
            if attachment:
                kwargs["attachments"] = attachment
            
            yag.send(**kwargs)
            logger.info(f"Email sent to {to_email}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
            return False
    
    def _get_overall_assessment(self, avg_score: float) -> str:
        """Generate overall assessment based on average score"""
        if avg_score >= 8.0:
            return "Kết quả bài thi rất tốt, đa số sinh viên nắm vững kiến thức."
        elif avg_score >= 6.5:
            return "Kết quả bài thi khá tốt, còn một số điểm cần cải thiện."
        elif avg_score >= 5.0:
            return "Kết quả ở mức trung bình, nhiều sinh viên còn hổng kiến thức."
        return "Kết quả thấp, cần xem lại đề thi hoặc phương pháp giảng dạy."


# Singleton instance
grading_service = GradingService()
