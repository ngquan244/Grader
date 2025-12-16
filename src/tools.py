"""
Tools definition for LangGraph Agent
Định nghĩa các tools theo format của LangChain
"""
import json
import sys
import datetime
import random
from pathlib import Path
from typing import Optional, Type
from langchain.tools import BaseTool
from pydantic import BaseModel, Field
from .notebook_tool import get_notebook_tool
from .config import Config
from .config import Config

receiver = Config.EMAIL_RECEIVER
user = Config.EMAIL_USER
password = Config.EMAIL_PASSWORD
import pyodbc
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
import yagmail
import logging

logger = logging.getLogger(__name__)


SQL_SERVER_CONN_STR = (
    "Driver={ODBC Driver 17 for SQL Server};"
    "Server=244-NGUYEN-QUAN\\SQL2022;"
    "Database=Agent;"
    "Trusted_Connection=yes;"
    "Encrypt=no;"
)

# Import quiz-gen utilities
sys.path.append(str(Path(__file__).parent.parent / "quiz-gen"))
try:
    from utils import extract_questions_from_pdf
except ImportError:
    extract_questions_from_pdf = None


# Tool Input Schema for Calculator
class CalculatorInput(BaseModel):
    """Input schema for calculator tool"""
    expression: str = Field(description="Biểu thức toán học cần tính, ví dụ: '2 + 2' hoặc '10 * 5'")


class CalculatorTool(BaseTool):
    """Tool để tính toán các phép toán đơn giản"""
    
    name: str = "calculator"
    description: str = """
    Công cụ tính toán toán học.
    Sử dụng khi người dùng muốn tính toán số học.
    Input: biểu thức toán học (string)
    Ví dụ: '2 + 2', '10 * 5 + 3', '100 / 4'
    """
    args_schema: Type[BaseModel] = CalculatorInput
    
    def _run(self, expression: str) -> str:
        """Execute calculator"""
        try:
            # Chỉ cho phép các ký tự an toàn
            allowed_chars = set("0123456789+-*/().% ")
            if not all(c in allowed_chars for c in expression):
                return json.dumps({
                    "error": "Biểu thức chứa ký tự không hợp lệ",
                    "allowed": "Chỉ được dùng: 0-9, +, -, *, /, (, ), %, space"
                }, ensure_ascii=False)
            
            result = eval(expression)
            return json.dumps({
                "expression": expression,
                "result": result
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({
                "error": str(e),
                "expression": expression
            }, ensure_ascii=False)
    
    async def _arun(self, expression: str) -> str:
        """Execute tool asynchronously"""
        return self._run(expression)


# Tool Input Schema for Quiz Generator
class QuizGeneratorInput(BaseModel):
    """Input schema for quiz generator tool"""
    num_questions: int = Field(
        default=10,
        description="Số lượng câu hỏi cần tạo cho quiz (mặc định 10)"
    )


class QuizGeneratorTool(BaseTool):
    """Tool để tự động tạo quiz từ file PDF trong thư mục data/quiz/"""
    
    name: str = "quiz_generator"
    description: str = """
    Công cụ tạo bài kiểm tra trực tuyến (quiz) từ đề thi PDF có sẵn.
    
    Chức năng:
    - Tự động đọc file PDF đề thi từ thư mục data/quiz/
    - Trích xuất các câu hỏi trắc nghiệm
    - Tạo file HTML standalone cho sinh viên làm bài
    - Trả về đường dẫn file HTML để chia sẻ
    
    Input: số lượng câu hỏi muốn tạo (mặc định 10)
    
    Sử dụng khi người dùng yêu cầu:
    - "Tạo quiz", "Tạo đề thi trắc nghiệm"
    - "Tạo bài kiểm tra online"
    - "Gen quiz từ PDF"
    - "Tạo quiz 15 câu", "Tạo quiz 20 câu"
    """
    args_schema: Type[BaseModel] = QuizGeneratorInput
    
    def _run(self, num_questions: int = 10) -> str:
        """Execute quiz generator"""
        try:
            if extract_questions_from_pdf is None:
                return json.dumps({
                    "error": "Module quiz-gen chưa được cài đặt",
                    "status": "failed"
                }, ensure_ascii=False)
            
            # Find PDF file in data/quiz/
            quiz_folder = Config.PROJECT_ROOT / "data" / "quiz"
            if not quiz_folder.exists():
                return json.dumps({
                    "error": "Thư mục data/quiz/ không tồn tại",
                    "status": "failed"
                }, ensure_ascii=False)
            
            pdf_files = list(quiz_folder.glob("*.pdf"))
            if not pdf_files:
                return json.dumps({
                    "error": "Không tìm thấy file PDF trong data/quiz/",
                    "status": "failed"
                }, ensure_ascii=False)
            
            # Use the first PDF file found
            pdf_file = pdf_files[0]
            
            # Extract questions
            questions = extract_questions_from_pdf(str(pdf_file))
            if not questions:
                return json.dumps({
                    "error": "Không tìm thấy câu hỏi trong PDF",
                    "pdf_file": pdf_file.name,
                    "status": "failed"
                }, ensure_ascii=False)
            
            # Randomly select questions
            selected_questions = random.sample(questions, min(num_questions, len(questions)))
            
            # Create quiz data
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            quiz_id = f"quiz_{timestamp}"
            
            quiz_data = {
                "id": quiz_id,
                "timestamp": timestamp,
                "source_pdf": pdf_file.name,
                "questions": selected_questions,
                "num_questions": len(selected_questions)
            }
            
            # Save quiz JSON
            output_folder = Config.PROJECT_ROOT / "quiz-gen" / "generated_quizzes"
            output_folder.mkdir(parents=True, exist_ok=True)
            
            quiz_json_file = output_folder / f"{quiz_id}.json"
            with open(quiz_json_file, 'w', encoding='utf-8') as f:
                json.dump(quiz_data, f, ensure_ascii=False, indent=2)
            
            # Generate HTML file
            html_file = output_folder / f"{quiz_id}.html"
            html_content = self._generate_html(quiz_id, selected_questions)
            
            with open(html_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Return success with file path
            file_url = f"file:///{str(html_file).replace(chr(92), '/')}"
            
            # Create a simple summary message
            summary = f"""
 ĐÃ TẠO QUIZ THÀNH CÔNG!

 Thông tin quiz:
   - Quiz ID: {quiz_id}
   - Nguồn: {pdf_file.name}
   - Số câu hỏi: {len(selected_questions)}/{len(questions)} câu

 File đã tạo:
   - HTML: {html_file.name}
   - JSON: {quiz_json_file.name}

 LINK QUIZ (Copy và dán vào trình duyệt):
   {file_url}

 Cách chia sẻ:
   1. Copy link trên và gửi cho sinh viên
   2. Hoặc gửi file: {html_file}
   3. Sinh viên mở bằng trình duyệt bất kỳ
   
 Lưu ý: File HTML hoạt động offline, không cần internet!
"""
            
            return json.dumps({
                "status": "success",
                "quiz_id": quiz_id,
                "source_pdf": pdf_file.name,
                "num_questions": len(selected_questions),
                "total_available": len(questions),
                "html_file": str(html_file),
                "file_url": file_url,
                "summary": summary,
                "message": summary
            }, ensure_ascii=False, indent=2)
            
        except Exception as e:
            return json.dumps({
                "error": str(e),
                "status": "failed"
            }, ensure_ascii=False)
    
    def _generate_html(self, quiz_id: str, questions: list) -> str:
        """Generate standalone HTML quiz file"""
        html = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title> {quiz_id}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Segoe UI', Arial, sans-serif; 
               background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
               padding: 20px; min-height: 100vh; }}
        .container {{ max-width: 900px; margin: 0 auto; background: white; 
                     border-radius: 20px; box-shadow: 0 20px 60px rgba(0,0,0,0.3); }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                  color: white; padding: 40px; text-align: center; border-radius: 20px 20px 0 0; }}
        .header h1 {{ font-size: 36px; margin-bottom: 10px; }}
        .content {{ padding: 40px; }}
        .student-form {{ background: #f8f9fa; padding: 30px; border-radius: 15px; margin-bottom: 30px; }}
        .form-group {{ margin-bottom: 20px; }}
        .form-group label {{ display: block; font-weight: bold; margin-bottom: 8px; 
                            color: #333; font-size: 16px; }}
        .form-group input {{ width: 100%; padding: 14px; border: 2px solid #ddd; 
                            border-radius: 10px; font-size: 16px; transition: all 0.3s; }}
        .form-group input:focus {{ border-color: #667eea; outline: none; box-shadow: 0 0 0 3px rgba(102,126,234,0.1); }}
        .question-card {{ background: #fff; border: 2px solid #e0e0e0; border-radius: 15px; 
                         padding: 30px; margin-bottom: 25px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        .question-text {{ font-size: 20px; font-weight: bold; color: #333; margin-bottom: 20px; }}
        .option-label {{ display: block; padding: 15px 20px; margin-bottom: 12px; 
                        border: 2px solid #ddd; border-radius: 10px; cursor: pointer; 
                        transition: all 0.3s; font-size: 16px; }}
        .option-label:hover {{ background: #f0f0f0; border-color: #667eea; transform: translateX(5px); }}
        .option-label input {{ margin-right: 12px; cursor: pointer; }}
        .result {{ margin-top: 20px; padding: 15px; border-radius: 10px; font-weight: bold; 
                  font-size: 16px; display: none; }}
        .score-panel {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                       color: white; padding: 30px; border-radius: 15px; text-align: center; 
                       margin: 30px 0; display: none; }}
        .score-panel h2 {{ font-size: 48px; margin: 20px 0; }}
        .submit-btn {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                      color: white; padding: 18px 50px; font-size: 20px; border: none; 
                      border-radius: 12px; cursor: pointer; display: block; margin: 40px auto; 
                      font-weight: bold; transition: all 0.3s; box-shadow: 0 4px 15px rgba(102,126,234,0.4); }}
        .submit-btn:hover {{ transform: scale(1.05); box-shadow: 0 6px 20px rgba(102,126,234,0.6); }}
        .info-display {{ background: #e3f2fd; padding: 20px; border-radius: 10px; 
                        margin-bottom: 30px; display: none; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1> Bài Kiểm Tra Trực Tuyến</h1>
            <p style="font-size: 18px; margin-top: 10px;">Quiz ID: {quiz_id}</p>
        </div>
        <div class="content">
            <div class="student-form" id="studentForm">
                <h3 style="margin-bottom: 25px; color: #667eea; font-size: 24px;"> Thông Tin Sinh Viên</h3>
                <div class="form-group">
                    <label>Họ và Tên *</label>
                    <input type="text" id="studentName" placeholder="Nhập họ tên đầy đủ" required>
                </div>
                <div class="form-group">
                    <label>Mã Sinh Viên *</label>
                    <input type="text" id="studentId" placeholder="Nhập mã sinh viên" required>
                </div>
                <div class="form-group">
                    <label>Email *</label>
                    <input type="email" id="studentEmail" placeholder="example@email.com" required>
                </div>
                <button onclick="startQuiz()" class="submit-btn"> Bắt Đầu Làm Bài</button>
            </div>
            
            <div class="info-display" id="infoDisplay"></div>
            
            <div id="quizContent" style="display: none;">
                <h3 style="color: #667eea; margin-bottom: 30px; font-size: 24px;"> Câu Hỏi ({len(questions)} câu)</h3>
"""
        
        # Add questions
        for i, q in enumerate(questions, start=1):
            html += f"""
                <div class="question-card">
                    <div class="question-text">Câu {i}: {q['question']}</div>
"""
            for letter, text in q['options'].items():
                html += f"""
                    <label class="option-label">
                        <input type="radio" name="q{i}" value="{letter}" 
                               data-correct="{q['correct']['letter']}" 
                               onclick="checkAnswer(this, {i})">
                        {letter}. {text}
                    </label>
"""
            html += f"""
                    <div id="result{i}" class="result"></div>
                </div>
"""
        
        html += f"""
                <button onclick="submitQuiz()" class="submit-btn"> Nộp Bài</button>
                <div class="score-panel" id="scorePanel">
                    <h3> KẾT QUẢ</h3>
                    <h2 id="scoreText"></h2>
                    <p id="scoreDetail" style="font-size: 18px; margin-top: 10px;"></p>
                </div>
            </div>
        </div>
    </div>
    
    <script>
        let studentInfo = {{}};
        let answers = {{}};
        const totalQuestions = {len(questions)};
        
        function startQuiz() {{
            const name = document.getElementById('studentName').value.trim();
            const id = document.getElementById('studentId').value.trim();
            const email = document.getElementById('studentEmail').value.trim();
            
            if (!name || !id || !email) {{
                alert(' Vui lòng điền đầy đủ thông tin!');
                return;
            }}
            
            studentInfo = {{ name, id, email }};
            
            document.getElementById('studentForm').style.display = 'none';
            document.getElementById('infoDisplay').style.display = 'block';
            document.getElementById('infoDisplay').innerHTML = `
                <h4 style="margin-bottom: 15px; color: #667eea;">👤 Thông tin của bạn:</h4>
                <p><strong>Họ tên:</strong> ${{name}}</p>
                <p><strong>MSSV:</strong> ${{id}}</p>
                <p><strong>Email:</strong> ${{email}}</p>
            `;
            document.getElementById('quizContent').style.display = 'block';
            
            window.scrollTo({{ top: 0, behavior: 'smooth' }});
        }}
        
        function checkAnswer(radio, questionNum) {{
            const correct = radio.getAttribute('data-correct');
            const result = document.getElementById('result' + questionNum);
            const labels = radio.parentNode.parentNode.querySelectorAll('.option-label');
            labels.forEach(l => {{
                l.style.background = '';
                l.style.borderColor = '#ddd';
            }});
            // Không hiển thị kết quả đúng/sai cho học sinh
            result.style.display = 'none';
            if(radio.value === correct) {{
                radio.parentNode.style.background = '#d4edda';
                radio.parentNode.style.borderColor = '#28a745';
                answers[questionNum] = true;
            }} else {{
                radio.parentNode.style.background = '#f8d7da';
                radio.parentNode.style.borderColor = '#dc3545';
                answers[questionNum] = false;
            }}
        }}
        
        function submitQuiz() {{
            let correctCount = 0;
            let answeredCount = 0;
            // Duyệt qua tất cả các câu hỏi
            for (let i = 1; i <= totalQuestions; i++) {{
                const radios = document.getElementsByName('q' + i);
                let selected = null;
                let correct = null;
                for (let r of radios) {{
                    if (r.checked) selected = r.value;
                    correct = r.getAttribute('data-correct');
                }}
                if (selected !== null) {{
                    answeredCount++;
                    if (selected === correct) correctCount++;
                }}
            }}
            if (answeredCount < totalQuestions) {{
                if (!confirm(`Bạn mới trả lời ${{answeredCount}}/${{totalQuestions}} câu. Bạn có chắc muốn nộp bài?`)) {{
                    return;
                }}
            }}
            // Tính điểm trên thang 10
            const score = ((correctCount / totalQuestions) * 10).toFixed(1);
            document.getElementById('scoreText').textContent = score + ' điểm';
            document.getElementById('scoreDetail').textContent = 
                `Đúng ${{correctCount}}/${{totalQuestions}} câu`;
            document.getElementById('scorePanel').style.display = 'block';
            window.scrollTo({{ top: document.getElementById('scorePanel').offsetTop - 100, behavior: 'smooth' }});
            console.log('Kết quả:', {{
                ...studentInfo,
                score: score,
                correct: correctCount,
                total: totalQuestions,
                quizId: '{quiz_id}'
            }});
        }}
    </script>
</body>
</html>"""
        
        return html
    
    async def _arun(self, num_questions: int = 10) -> str:
        """Execute tool asynchronously"""
        return self._run(num_questions)

from typing import Type, List, Dict, Any
from langchain.tools import BaseTool
from pydantic import BaseModel, Field
import pyodbc
import json
from .logger import logger

SQL_SERVER_CONN_STR = (
    "Driver={ODBC Driver 17 for SQL Server};"
    "Server=244-NGUYEN-QUAN\\SQL2022;"
    "Database=Agent;"
    "Trusted_Connection=yes;"
    "Encrypt=no;"
)


class ExamSummaryInput(BaseModel):
    exam_code: str = Field(
        description="Mã đề thi cần tổng hợp (ví dụ: 000)"
    )



class ExamResultSummaryTool(BaseTool):
    """
    Tool tổng hợp kết quả bài thi theo mã đề
    """

    name: str = "summarize_exam_results"
    description: str = """
    Tổng hợp kết quả chấm điểm từ bảng FinalExamResult theo mã đề thi.

    Sử dụng khi người dùng yêu cầu:
    - Tổng hợp kết quả mã đề
    - Thống kê điểm bài thi
    - Xem kết quả toàn bộ sinh viên theo mã đề

    Tool sẽ:
    1. Query các cột cần thiết:
       - student_id, name, email, exam_code, score
    2. Trả về:
       - Danh sách sinh viên
       - Điểm trung bình, cao nhất, thấp nhất
       - Đánh giá tổng quan kết quả bài thi
       - File Excel tổng hợp
       - Gửi email file Excel đến địa chỉ cố định
    """

    args_schema: Type[BaseModel] = ExamSummaryInput


    def _run(self, exam_code: str) -> str:
        try:
            logger.info(f" Summarizing results for exam_code={exam_code}")

            conn = pyodbc.connect(SQL_SERVER_CONN_STR)
            cursor = conn.cursor()

            sql = """
            SELECT
                student_id,
                name,
                email,
                exam_code,
                score
            FROM dbo.FinalExamResult
            WHERE exam_code = ?
            ORDER BY score DESC
            """

            cursor.execute(sql, exam_code)
            rows = cursor.fetchall()

            if not rows:
                return json.dumps({
                    "exam_code": exam_code,
                    "message": "Không tìm thấy kết quả cho mã đề này"
                }, ensure_ascii=False, indent=2)

            results: List[Dict[str, Any]] = []
            scores: List[float] = []

            for r in rows:
                score = float(r.score)
                scores.append(score)

                results.append({
                    "student_id": r.student_id,
                    "name": r.name,
                    "email": r.email,
                    "exam_code": r.exam_code,
                    "score": score,
                    "evaluation": self._evaluate_score(score)
                })

            summary = {
                "total_students": len(scores),
                "average_score": round(sum(scores) / len(scores), 2),
                "max_score": max(scores),
                "min_score": min(scores),
            }

            assessment = self._overall_assessment(summary["average_score"])

            excel_file = self._export_to_excel(
                exam_code=exam_code,
                summary=summary,
                results=results
            )

            # Gửi email cố định
            subject = f"Kết quả tổng hợp mã đề {exam_code}"
            body = f"Đính kèm file Excel tổng hợp kết quả bài thi mã đề {exam_code}."
            sent = self._send_excel_email(
                file_path=excel_file,
                to_email=Config.EMAIL_RECEIVER,
                subject=subject,
                body=body
            )
            if sent:
                logger.info(f"File Excel đã được gửi đến {Config.EMAIL_RECEIVER}")
            else:
                logger.warning(f"Không thể gửi file Excel đến {Config.EMAIL_RECEIVER}")

            
            cursor.close()
            conn.close()

            return json.dumps({
                "exam_code": exam_code,
                "summary": summary,
                "overall_assessment": assessment,
                "results": results,
                "excel_file": excel_file
            }, ensure_ascii=False, indent=2)

        except Exception as e:
            logger.error(f" Error summarizing exam results: {str(e)}")
            logger.exception(e)
            return json.dumps({
                "error": str(e),
                "type": type(e).__name__
            }, ensure_ascii=False, indent=2)

    async def _arun(self, exam_code: str) -> str:
        return self._run(exam_code)

    def _evaluate_score(self, score: float) -> str:
        if score >= 8.5:
            return "Xuất sắc"
        elif score >= 7.0:
            return "Tốt"
        elif score >= 5.0:
            return "Đạt"
        else:
            return "Chưa đạt"

    def _overall_assessment(self, avg_score: float) -> str:
        if avg_score >= 8.0:
            return "Kết quả bài thi rất tốt, đa số sinh viên nắm vững kiến thức."
        elif avg_score >= 6.5:
            return "Kết quả bài thi khá tốt, còn một số điểm cần cải thiện."
        elif avg_score >= 5.0:
            return "Kết quả ở mức trung bình, nhiều sinh viên còn hổng kiến thức."
        else:
            return "Kết quả thấp, cần xem lại đề thi hoặc phương pháp giảng dạy."
    
    def _send_excel_email(self, file_path: str, to_email: str, subject: str, body: str):
        """
        Gửi file Excel qua email
        """
        try:
            # Config tài khoản gửi email (có thể lưu ở Config)
            yag = yagmail.SMTP(user=Config.EMAIL_USER, password=Config.EMAIL_PASSWORD)
            yag.send(
                to=to_email,
                subject=subject,
                contents=body,
                attachments=file_path
            )
            logger.info(f"Excel file sent to {to_email}")
            return True
        except Exception as e:
            logger.error(f"Failed to send email: {str(e)}")
            return False    

    def _export_to_excel(
        self,
        exam_code: str,
        summary: Dict[str, Any],
        results: List[Dict[str, Any]]
    ) -> str:
        wb = Workbook()

        # ===== Sheet 1: Summary =====
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary["A1"] = "Exam Code"
        ws_summary["B1"] = exam_code

        ws_summary["A3"] = "Total Students"
        ws_summary["B3"] = summary["total_students"]

        ws_summary["A4"] = "Average Score"
        ws_summary["B4"] = summary["average_score"]

        ws_summary["A5"] = "Max Score"
        ws_summary["B5"] = summary["max_score"]

        ws_summary["A6"] = "Min Score"
        ws_summary["B6"] = summary["min_score"]

        for cell in ["A1", "A3", "A4", "A5", "A6"]:
            ws_summary[cell].font = Font(bold=True)

        # ===== Sheet 2: Results =====
        ws_results = wb.create_sheet("Results")

        headers = [
            "Student ID",
            "Name",
            "Email",
            "Exam Code",
            "Score",
            "Evaluation"
        ]
        ws_results.append(headers)

        for col in range(1, len(headers) + 1):
            ws_results.cell(row=1, column=col).font = Font(bold=True)

        for r in results:
            ws_results.append([
                r["student_id"],
                r["name"],
                r["email"],
                r["exam_code"],
                r["score"],
                r["evaluation"]
            ])

        for col in ws_results.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_results.column_dimensions[col[0].column_letter].width = max_len + 2

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"exam_summary_{exam_code}_{timestamp}.xlsx"

        output_dir = Config.PROJECT_ROOT / "exports"
        output_dir.mkdir(parents=True, exist_ok=True)

        file_path = output_dir / filename
        wb.save(file_path)

        return str(file_path)



# Registry of all available tools
def get_all_tools() -> list[BaseTool]:
    """Trả về danh sách tất cả các tools có sẵn"""
    return [
        get_notebook_tool(),
        CalculatorTool(),
        QuizGeneratorTool(),
        ExamResultSummaryTool()
    ]


def get_tool_by_name(tool_name: str) -> Optional[BaseTool]:
    """Lấy tool theo tên"""
    tools = get_all_tools()
    for tool in tools:
        if tool.name == tool_name:
            return tool
    return None
